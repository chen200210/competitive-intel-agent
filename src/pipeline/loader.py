"""
Data Loader — parse raw ranking files (CSV / Excel) into normalized records.

Supports:
- .csv (utf-8-sig encoding, header in row 1)
- .xlsx (openpyxl, header may be in row 2 with metadata in row 1)

Handles data quality issues:
- bundle_id may be "0" → fallback to "fallback:{game_name}"
- developer may be "0" → converted to None
- Date extracted from filename: "20260616" or "2026-06-16"
- chart_type extracted from filename: "热门榜" / "免费榜" / "畅销榜" etc.
- Platform normalized: "IOS游戏榜" → "iOS", "ANDROID游戏榜" → "Android"

Usage:
    python -m src.pipeline.loader --file data/raw/TapTap_xxx_20260616.xlsx
"""

import csv
import re
from pathlib import Path
from typing import Any

import openpyxl

from src.storage.sqlite import get_db


# ── Field mapping ────────────────────────────────────────────

# Known column headers (after stripping whitespace) → internal field name
COLUMN_ALIASES: dict[str, str] = {
    # TapTap / 七麦 中文列名
    "排名": "rank",
    "Bundle ID": "bundle_id",
    "应用": "game_name",
    "Android/iOS": "platform_raw",
    "iOS/Android": "platform_raw",
    "类别": "category",
    "开发者": "developer",
    # 点点数据 / 英文列名
    "platform": "platform_raw",
    "scrape_date": "date_override",
    "genre": "category",
    "publisher": "developer",
}


def normalize_platform(raw: str) -> str:
    """Normalize platform field. 'IOS游戏榜' → 'iOS', etc."""
    raw = raw.strip().upper()
    if "IOS" in raw:
        return "iOS"
    if "ANDROID" in raw:
        return "Android"
    return raw


def extract_date_from_filename(file_path: str | Path) -> str:
    """
    Extract YYYY-MM-DD date from filename.

    Supports:
      - "2026-06-16_rankings.csv"  (hyphenated)
      - "TapTap_xxx_20260616.xlsx" (compact, 8 digits)
    """
    name = Path(file_path).name
    # Try hyphenated first: 2026-06-16
    match = re.search(r"(\d{4}-\d{2}-\d{2})", name)
    if match:
        return match.group(1)
    # Try compact 8-digit: 20260616
    match = re.search(r"(\d{4})(\d{2})(\d{2})", name)
    if match:
        return f"{match.group(1)}-{match.group(2)}-{match.group(3)}"
    raise ValueError(f"Cannot extract date from filename: {name}")


def _detect_header_row(ws) -> int:
    """
    Detect which row contains column headers.

    Returns 0-based row index. For TapTap Excel files, headers are
    usually in row 2 (index 1) while row 1 is metadata.
    """
    # Check first 3 rows for known column headers
    for row_idx in range(min(3, ws.max_row)):
        cell_values = [
            str(ws.cell(row_idx + 1, col + 1).value or "").strip()
            for col in range(ws.max_column)
        ]
        matches = sum(1 for v in cell_values if v in COLUMN_ALIASES)
        if matches >= 4:  # At least 4 known columns found
            return row_idx
    return 0  # Fallback: assume first row


def extract_chart_type_from_filename(file_path: str | Path) -> str:
    """
    Extract chart type from filename.

    Supports:
      - "TapTap_Android游戏榜_热门榜_20260616.xlsx" → "热门榜"
      - "ios_game_free_rank_20260616.csv"        → "免费榜"
      - "android_top_grossing_20260616.csv"      → "畅销榜"

    Returns "热门榜" as default if no known keyword is found.
    Returns None if the file is NOT a ranking CSV (news, new games, etc.).
    """
    name = Path(file_path).name

    # Non-ranking files from other scrapers — skip these
    non_ranking_markers = ["资讯", "Steam移植", "新游", "B站", "bilibili", "steam_port"]
    for marker in non_ranking_markers:
        if marker in name:
            return None

    # Known chart type keywords (order matters — longer match first)
    keywords = ["热门榜", "免费榜", "畅销榜", "新品榜", "下载榜", "收入榜",
                "免费", "畅销", "热门", "新品", "下载", "收入"]

    for kw in keywords:
        if kw in name:
            # Normalize to standard form
            if kw in ("免费", "免费榜"):
                return "免费榜"
            if kw in ("畅销", "畅销榜"):
                return "畅销榜"
            if kw in ("热门", "热门榜"):
                return "热门榜"
            if kw in ("新品", "新品榜"):
                return "新品榜"
            if kw in ("下载", "下载榜"):
                return "下载榜"
            if kw in ("收入", "收入榜"):
                return "收入榜"

    return "热门榜"  # default


def parse_row(
    raw_values: dict[str, str], date: str, chart_type: str, source_file: str
) -> dict[str, Any]:
    """Parse a single row (already column-mapped) into the internal record format."""
    game_name = raw_values.get("game_name", "").strip()
    bundle_id_raw = raw_values.get("bundle_id", "").strip()
    developer_raw = raw_values.get("developer", "").strip()
    platform_raw = raw_values.get("platform_raw", "").strip()
    category = raw_values.get("category", "").strip()
    date_override = raw_values.get("date_override", "").strip()

    # Use row-level date if present (e.g. diandian scrape_date column), else file-level date
    effective_date = date_override if date_override else date

    # Data quality: bundle_id may be "0" or empty
    if bundle_id_raw == "0" or not bundle_id_raw:
        bundle_id = f"fallback:{game_name}"
    else:
        bundle_id = bundle_id_raw

    # Data quality: developer may be "0"
    developer = None if developer_raw == "0" else developer_raw

    rank_str = raw_values.get("rank", "0").strip()
    rank = int(rank_str) if rank_str.isdigit() else 0

    return {
        "date": effective_date,
        "platform": normalize_platform(platform_raw),
        "chart_type": chart_type,
        "category": category,
        "rank": rank,
        "bundle_id": bundle_id,
        "game_name": game_name,
        "developer": developer,
        "source_file": source_file,
    }


# ── File readers ─────────────────────────────────────────────

def _read_xlsx(file_path: Path, date: str, chart_type: str) -> list[dict[str, Any]]:
    """Read ranking data from an Excel file."""
    wb = openpyxl.load_workbook(file_path, read_only=True)
    ws = wb.active

    header_row = _detect_header_row(ws)
    headers = [
        str(ws.cell(header_row + 1, col + 1).value or "").strip()
        for col in range(ws.max_column)
    ]

    # Map column indices to internal field names
    col_map: dict[int, str] = {}
    for idx, h in enumerate(headers):
        if h in COLUMN_ALIASES:
            col_map[idx] = COLUMN_ALIASES[h]

    records: list[dict[str, Any]] = []
    for row_idx in range(header_row + 1, ws.max_row):
        raw_values: dict[str, str] = {}
        for col_idx, field_name in col_map.items():
            cell_val = ws.cell(row_idx + 1, col_idx + 1).value
            raw_values[field_name] = str(cell_val).strip() if cell_val is not None else ""

        if not raw_values.get("game_name"):
            continue  # skip empty rows

        try:
            record = parse_row(raw_values, date, chart_type, file_path.name)
            records.append(record)
        except (ValueError, KeyError) as e:
            print(f"  [WARN] Skipping row {row_idx + 1}: {e}")

    wb.close()
    return records


def _read_csv(file_path: Path, date: str, chart_type: str) -> list[dict[str, Any]]:
    """Read ranking data from a CSV file."""
    records: list[dict[str, Any]] = []

    with open(file_path, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            # Remap column names through aliases
            raw_values: dict[str, str] = {}
            for csv_col, val in row.items():
                col_clean = csv_col.strip()
                field_name = COLUMN_ALIASES.get(col_clean, col_clean)
                raw_values[field_name] = val

            if not raw_values.get("game_name"):
                continue

            try:
                record = parse_row(raw_values, date, chart_type, file_path.name)
                records.append(record)
            except (ValueError, KeyError) as e:
                print(f"  [WARN] Skipping row: {e}")

    return records


# ── Main entry point ─────────────────────────────────────────

def import_file(
    file_path: str | Path,
    date: str | None = None,
    chart_type: str | None = None,
) -> dict[str, Any]:
    """
    Import a ranking data file (CSV or Excel) into the rankings table.

    Args:
        file_path: Path to .csv or .xlsx file.
        date: Date string YYYY-MM-DD. Auto-extracted from filename if None.
        chart_type: Chart type (热门榜/免费榜/...). Auto-extracted if None.

    Returns:
        Summary dict with count of imported records.
    """
    file_path = Path(file_path)

    if date is None:
        date = extract_date_from_filename(file_path)

    if chart_type is None:
        chart_type = extract_chart_type_from_filename(file_path)
        if chart_type is None:
            return {"imported": 0, "date": date, "skipped": True,
                    "reason": "not a ranking file"}

    suffix = file_path.suffix.lower()
    if suffix == ".xlsx":
        records = _read_xlsx(file_path, date, chart_type)
    elif suffix == ".csv":
        records = _read_csv(file_path, date, chart_type)
    else:
        return {"imported": 0, "date": date, "error": f"Unsupported file type: {suffix}"}

    if not records:
        return {"imported": 0, "date": date, "error": "No valid records found"}

    db = get_db()
    db.insert_rankings(records)

    return {
        "imported": len(records),
        "date": date,
        "chart_type": chart_type,
        "source": file_path.name,
    }


# Alias for backward compatibility
import_csv = import_file


# ── CLI test entry ───────────────────────────────────────────

if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2 or sys.argv[1] != "--file":
        print("Usage: python -m src.pipeline.loader --file data/raw/TapTap_xxx.xlsx")
        sys.exit(1)

    result = import_file(sys.argv[2])
    print(result)
